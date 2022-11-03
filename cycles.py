# -*- coding: utf-8 -*-
"""
Created on Thu Aug 22 14:25:59 2019

@author: wiktor-jedski

Flight hours/cycles/days controls calculator for parts installed on A/C.
Script is based on Excel reports taken from CAMO software and returns a report with total hours, cycles and days for
specific controls for a given component.

A/C report column names:
AC | FLIGHT_DAT | OFF_HOUR | ON_HOUR | FLIGHT_HOURS | FLIGHT_MINS | CYCLES | TOTAL_AC_FLIGHT_HOUR | TOTAL_AC_FLIGHT_MIN
| TOTAL_AC_CYCLES | FLIGHT_LOG
P/N report column names:
PN | SN | CONTROL | SCHEDULE_HOURS | SCHEDULE_CYCLES | SCHEDULE_DAYS | ACTUAL_HOURS | ACTUAL_CYCLES | ACTUAL_DAYS
| INSTALLED_ON | INSTALLED_DATE
Output report column names:
PN | SN | CONTROL | SCHEDULE_HOURS | SCHEDULE_CYCLES | SCHEDULE_DAYS | ACTUAL_HOURS | ACTUAL_CYCLES | ACTUAL_DAYS
| INSTALLED_ON | INSTALLED_DATE | TOTAL_HRS | TOTAL_CYC | TOTAL_DAYS
"""

import sys
from openpyxl import *
from datetime import datetime

# read file name from command line argument and load
print('Loading components file...')
try:
    file = sys.argv[1]
except IndexError:
    print("Usage: python3 cycles.py report_components.xlsx report_ac.xlsx")
    quit()
wb_components = load_workbook(filename=file, data_only=True)
ws_comp = wb_components.active
print(str(ws_comp.max_row - 1) + ' controls loaded.')

print('Loading A/C flights file...')
try:
    file = sys.argv[2]
except IndexError:
    print("Usage: python3 cycles.py report_components.xlsx report_ac.xlsx")
    quit()
wb_ac = load_workbook(filename=file, data_only=True)
ws_ac = wb_ac.active 
print(str(ws_ac.max_row - 1) + ' flight logs loaded.')

# read hours anc cycles of an ac from the report_ac file
ac_hours = ws_ac['H' + str(ws_ac.max_row)].value
ac_cycles = ws_ac['J' + str(ws_ac.max_row)].value

# formatting of the output file
ws_comp['L1'] = "TOTAL_HRS"
ws_comp['M1'] = "TOTAL_CYC"
ws_comp['N1'] = "TOTAL_DAYS"

# main loop
for row in range(2, ws_comp.max_row + 1):
    date_installed = ws_comp['K' + str(row)].value
    # if part has no installation date or if part installation date is smaller than z date,
    # take z date
    if date_installed is None or date_installed < ws_ac['B2'].value:
        date_installed = ws_ac['B2'].value
        
    # find the first date larger than itself
    for line in range(2, ws_ac.max_row + 1):
        if ws_ac['B' + str(line)].value >= date_installed:
            hrs_before = ws_ac['H' + str(line - 1)].value
            cyc_before = ws_ac['J' + str(line - 1)].value
            break
    # date not found - assume 0 before installation
    else:
        hrs_before = 0
        cyc_before = 0
        
    # calculate difference between ac_hours, ac_cycles and before
    hrs_now = ws_comp['G' + str(row)].value + (ac_hours - int(hrs_before))
    cyc_now = ws_comp['H' + str(row)].value + (ac_cycles - int(cyc_before))
    days_delta = datetime.now() - date_installed
    days_now = ws_comp['I' + str(row)].value + days_delta.days
    
    # write in new data
    ws_comp['L' + str(row)] = hrs_now
    ws_comp['M' + str(row)] = cyc_now
    ws_comp['N' + str(row)] = days_now
    
    # give feedback
    print('PN: ' + str(ws_comp['A' + str(row)].value) + ' SN: ' + str(ws_comp['B' + str(row)].value) + ' Control: ' +
          str(ws_comp['C' + str(row)].value) + ' has been computed.')
    
wb_components.save(filename='output_' + str(ws_ac['A2'].value) + '.xlsx')
print('Done.')
